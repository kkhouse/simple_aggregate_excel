import csv
import openpyxl
import sys
import os
import openpyxl as px


aggregate_list =["遠田郡","塩釜市","塩竈市","牡鹿郡","加美郡","刈田郡","岩沼市","七ヶ浜町","利府町","栗原市","黒川郡","柴田郡","石巻市","宮城野区","若林区","青葉区","泉区","太日区","多賀城","大崎市","登米市","東松島市","富谷市","本吉郡","名取市","亘理郡","その他"]

# ~~csv→xlsx変換 start~~~
wb = openpyxl.Workbook()
ws = wb.active
query_word = sys.argv[1] 

f = open("./data_csv/{}.csv".format(query_word))
reader = csv.reader(f, delimiter=":")
for row in reader:
  ws.append(row)
f.close()

if not os.path.exists("./data_xlsx"):
  os.makedirs("./data_xlsx")
wb.save("./data_xlsx/{0}{1}.xlsx".format(query_word,"_集計"))
# ~~csv→xlsx convert end~~~

# ~~xlsx Extraction start~~
folder = "data_xlsx"

filePath = []
for root,dirs,files in os.walk(folder):
  for fname in files:
    fnamePath = os.path.join(root,fname)
    if fnamePath.find(".xlsx")!= -1:
      filePath.append(fnamePath)
    if filePath ==[]:
      print("データがありません!")
xlsx_files_list = [s for s in filePath if query_word in s]



#作成したxlsxの取得
wb = px.load_workbook(xlsx_files_list[0])
ws = wb[wb.sheetnames[0]]

#xlsx内の住所のすべてをリスト化
adress_list = []

table = [0] * len(aggregate_list)

for i in range(ws.max_row):
  temp = ws.cell(row=i+1, column=1).value # 何回も出てくるので
  adress_list.append(temp)

  for j, name in enumerate(aggregate_list):
    if name in temp:
      table[j] += 1 # 対応するところへカウント
      break
  else:
    table[26] += 1 # その他

#塩釜と塩竃をまとめる
table[1] = table[1]+table[2]
table.pop(2)


for i,j in zip(list(range(1,3)),aggregate_list):
  ws.cell(row=i, column=7).value = j
for i,j in zip(list(range(3,27)),aggregate_list[3:]):
  ws.cell(row=i, column=7).value = j
for i,g in zip(list(range(1,28)),table):
  ws.cell(row=i, column=8).value = g

wb.save("./data_xlsx/{0}{1}.xlsx".format(query_word,"_集計"))
